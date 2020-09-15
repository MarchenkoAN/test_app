from django.views import View
from qwestion.models import Rating
from .filter import RatintSurveyFilter
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django_filters.views import FilterView
from django.db.models import Count


# Create your views here.

class RatingCountListView(FilterView):
    """
    Подсчет правльных ответов
    """
    breadcrumbs = ['ratiglist']
    model = Rating
    paginate_by = 25
    filterset_class = RatintSurveyFilter


    template_name = 'report/rating_list.html'

    def get_context_data(self, **kwargs):
        """
        Забираем параметры фильтра из строки запроса
        :param kwargs:
        :return:
        """
        context = super().get_context_data(**kwargs)
        survey = self.request.GET.get('survey','')
        survey = survey.replace(" ", "+")
        context['survey_id']=survey
        return context

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        """
        Декорируем диспетчер функцией login_required, чтобы запретить просмотр отображения неавторизованными
        пользователями
        """
        return super(RatingCountListView, self).dispatch(request, *args, **kwargs)

    def get_queryset(self):
        """

        Подсчитываем количество правильных ответов с учетом фильтра
        :return:
        """

        object_list = Rating.objects.values('sessionid','username').filter(result=True).order_by('sessionid').annotate(count=Count('id'))

        return object_list


class RatingToXls(View):
    """
    Формируем файл xls для модели Raiting
    """


    def get_context_data(self, **kwargs):
        """
        Забираем параметры фильтра из строки запроса
        :param kwargs:
        :return:
        """
        context = super().get_context_data(**kwargs)
        survey = self.request.GET.get('survey','')
        survey = survey.replace(" ", "+")
        context['survey_id']=survey
        return context


    @method_decorator(login_required)
    def get(self, request):
        ratings = Rating.objects.all()
        survey = self.request.GET.get('survey','')
        if (survey!=''):
            print(ratings)
            ratings = Rating.objects.filter(survey__pk=int(survey))
            print(ratings)
        response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        response['Content-Disposition'] = 'attachment; filename={date}-raing.xlsx'.format(
                date=datetime.now().strftime('%Y-%m-%d'),
            )
        workbook = Workbook()

            # Get active worksheet/tab
        worksheet = workbook.active

        worksheet.title = 'Результаты тестивания'

            # Define the titles for columns
        columns = [
                'ID',
                'Наименование теста',
                'Имя пользователя',
                'Email',
                'Вопрос',
                'Ответы пользователя',
                'Правильные ответы',
                'Результат',
                'Дата теста',
                'Сессия',
            ]
        row_num = 1

            # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

            # Iterate through all movies
        for rating in ratings:
            row_num += 1

                # Define the data for each cell in the row
            row = [
                    rating.pk,
                    str(rating.survey),
                    rating.username,
                    rating.email,
                    str(rating.qwestion),
                    rating.answer,
                    rating.true_answer,
                    rating.result,
                    rating.created,
                    rating.sessionid,
                ]

                # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response



